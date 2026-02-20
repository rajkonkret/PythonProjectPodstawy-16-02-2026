from openpyxl import load_workbook

wb = load_workbook('sales.xlsx')

ws = wb.active

print(ws['A1'].value)  # -> Sales Date

ws['A1'] = "Witaj"
ws['B2'] = 123

print(ws['A1'].value)  # -> Witaj

wb.save("plik.xlsx")
